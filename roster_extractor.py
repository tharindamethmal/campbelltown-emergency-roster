import openpyxl
from datetime import datetime
import pandas as pd

roster_file_path = 'Term 1, 2024 FINAL.xlsx'
sheet_name = 'Term 1 2024 - Under Review '
dr_name = 'dr name here as in the roster file'


def load_excel(roster_file_path, sheet_name):
    try:
        # Load the Excel file
        workbook = openpyxl.load_workbook(roster_file_path)
        # Check if the specified sheet exists
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        else:
            print(f"Sheet '{sheet_name}' not found in the Excel file.")
    except Exception as e:
        print(f"Error: {e}")


def search_text_in_sheet(workbook_sheet, target_text):
    try:

        # Iterate through all rows and columns in the specified sheet
        for row in workbook_sheet.iter_rows(min_row=1, max_row=workbook_sheet.max_row, min_col=1, max_col=workbook_sheet.max_column):
            for cell in row:
                # Check if the target text is in the cell
                if target_text.lower() in str(cell.value).lower():
                    return cell

    except Exception as e:
        print(f"Error: {e}")


def extract_values_from_row(workbook_sheet, row_number):
    try:

        # Retrieve values from the specified row
        return workbook_sheet[row_number]

    except Exception as e:
        print(f"Error: {e}")
        return None


work_sheet = load_excel(roster_file_path, sheet_name)
name_cell = search_text_in_sheet(work_sheet, dr_name)
date_cell = search_text_in_sheet(work_sheet, 'Date')


final_data = pd.DataFrame(columns=['Date', 'Shift'])


for date_cell in work_sheet[date_cell.row]:
    if isinstance(date_cell.value, datetime):
        shift = work_sheet.cell(
            row=name_cell.row, column=date_cell.column).value

        new_row = {'Date': date_cell.value.strftime(
            "%Y-%m-%d"), 'Shift': shift}
        final_data.loc[len(final_data)] = new_row

final_data.to_excel(dr_name+"_"+sheet_name+"_roster.xlsx")
